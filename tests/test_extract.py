"""extract の正規化・ファイル処理の単体テスト（API 呼び出しなし）。"""

import io

import pytest

from src.extract import (
    INDUSTRY_TYPES,
    _build_content_blocks,
    extract_excel_text,
    normalize_quote,
)


def _make_xlsx_bytes():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積書"
    ws["A1"] = "御見積書"
    ws["A3"] = "株式会社テスト工業 御中"
    ws["A5"] = "品名"
    ws["B5"] = "数量"
    ws["C5"] = "金額"
    ws["A6"] = "工場間通路テント"
    ws["B6"] = "1式"
    ws["C6"] = 1500000
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_excel_text_contains_cells():
    text = extract_excel_text(_make_xlsx_bytes())
    assert "御見積書" in text
    assert "工場間通路テント" in text
    assert "1500000" in text


def test_build_content_blocks_excel():
    blocks = _build_content_blocks("mitsumori.xlsx", _make_xlsx_bytes())
    assert blocks[0]["type"] == "text"
    assert "工場間通路テント" in blocks[0]["text"]


def test_build_content_blocks_image():
    blocks = _build_content_blocks("photo.PNG", b"\x89PNG fake")
    assert blocks[0]["type"] == "image"
    assert blocks[0]["source"]["media_type"] == "image/png"


def test_build_content_blocks_pdf():
    blocks = _build_content_blocks("quote.pdf", b"%PDF-1.4 fake")
    assert blocks[0]["type"] == "document"
    assert blocks[0]["source"]["media_type"] == "application/pdf"


def test_build_content_blocks_unsupported():
    with pytest.raises(ValueError):
        _build_content_blocks("readme.txt", b"hello")


def test_normalize_quote_fills_missing_fields():
    result = normalize_quote({})
    assert result["customer_name"] == ""
    assert result["items"] == []
    assert result["total_amount"] is None
    assert result["industry_type"] in INDUSTRY_TYPES


def test_normalize_quote_fixes_invalid_industry():
    result = normalize_quote({"industry_type": "銀行"})
    assert result["industry_type"] == "その他"


def test_normalize_quote_coerces_item_fields():
    result = normalize_quote(
        {"items": [{"name": "テント", "quantity": 1, "unit_price": None}]}
    )
    item = result["items"][0]
    assert item["name"] == "テント"
    assert item["quantity"] == "1"
    assert item["spec"] == ""
    assert item["amount"] is None
